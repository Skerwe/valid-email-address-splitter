import re
import openpyxl
from pathlib import Path

regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
original_filename = 'email_address_list.xlsx'
data_folder = 'data'


def run():
    print("\nStarting email address checker ...\n")

    xlsx_file = Path(data_folder, original_filename)
    wb_obj = openpyxl.load_workbook(xlsx_file)

    # Read the active sheet:
    sheet = wb_obj.active

    print("Number of rows:", sheet.max_row)
    print("Number of columns:", sheet.max_column)
    print()

    fixed_emails = []
    unable_to_fix = []

    for row in sheet.iter_rows():
        for cell in row:
            parts = re.compile(regex).findall(cell.value)
            if len(parts):
                fixed_emails.extend(parts)
            else:
                unable_to_fix.append(cell.value)

    if len(fixed_emails):
        save_workbook(fixed_emails)
    if len(unable_to_fix):
        save_workbook(unable_to_fix, 'require_manual_fix')

    print("Process complete.")


def save_workbook(email_list, prefix='fixed'):
    workbook = openpyxl.Workbook()

    dest_filename = f'{data_folder}/{prefix}_{original_filename}'

    sheet = workbook.active
    sheet.title = prefix

    for email in email_list:
        sheet.append([email])

    workbook.save(filename=dest_filename)
